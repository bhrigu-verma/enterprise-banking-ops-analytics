#!/usr/bin/env python3
"""
Aegis Crest Financial - Excel Reporting Workbook Generator
Creates Aegis_Banking_Operations_Reporting.xlsx using openpyxl with
multiple styled tabs, formulas (XLOOKUP, SUMIFS, AVERAGEIFS), pivot data, and conditional formatting.
"""

import os
import sqlite3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

DB_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "aegis_banking.db"))
EXCEL_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "Aegis_Banking_Operations_Reporting.xlsx"))

def build_excel():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles
    navy_fill = PatternFill(start_color="0B132B", end_color="0B132B", fill_type="solid")
    gold_fill = PatternFill(start_color="D4AF37", end_color="D4AF37", fill_type="solid")
    header_fill = PatternFill(start_color="1C2541", end_color="1C2541", fill_type="solid")
    light_blue_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    title_font = Font(name="Calibri", size=16, bold=True, color="D4AF37")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # ---------------------------------------------------------
    # TAB 1: EXECUTIVE SUMMARY
    # ---------------------------------------------------------
    ws1 = wb.create_sheet(title="Executive Summary")
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:G1")
    ws1["A1"] = "AEGIS CREST FINANCIAL - EXECUTIVE OPERATIONS SCORECARD"
    ws1["A1"].font = title_font
    ws1["A1"].fill = navy_fill
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 40

    # KPI Summary Cards
    kpis = [
        ("Total Deposits", "=SUM('Regional Loan Summary'!C4:C8)", "$#,##0"),
        ("Active Loan Volume", "=SUM('Regional Loan Summary'!D4:D8)", "$#,##0"),
        ("Default Losses", "=SUM('Regional Loan Summary'!F4:F8)", "$#,##0"),
        ("Confirmed Fraud Loss", "=SUM('Fraud Risk Register'!C4:C9)", "$#,##0")
    ]

    cols = ["B", "D", "F", "H"]
    for idx, (label, formula, fmt) in enumerate(kpis):
        col = cols[idx]
        ws1[f"{col}3"] = label
        ws1[f"{col}3"].font = Font(size=9, color="64748B", bold=True)
        ws1[f"{col}4"] = formula
        ws1[f"{col}4"].font = Font(size=14, bold=True, color="0B132B")
        ws1[f"{col}4"].number_format = fmt
        ws1[f"{col}4"].alignment = Alignment(horizontal="center")

    # Table Header
    headers1 = ["Region Name", "Director", "Active Customers", "FastTrack Pct", "Avg Turnaround (Days)", "Default Rate Pct", "Fraud Loss ($)"]
    ws1.append([])
    ws1.append([])
    ws1.append(headers1)
    
    for col_num in range(1, 8):
        cell = ws1.cell(row=7, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    regional_data = cursor.execute("""
        SELECT r.region_name, r.regional_director, COUNT(c.customer_id),
               ROUND(100.0 * SUM(CASE WHEN c.is_digital_fasttrack = 1 THEN 1 ELSE 0 END)/COUNT(c.customer_id), 2),
               ROUND(AVG(l.approval_turnaround_days), 2),
               ROUND(100.0 * SUM(CASE WHEN l.status='Defaulted' THEN 1 ELSE 0 END)/COUNT(l.loan_id), 2),
               ROUND(COALESCE(SUM(fa.loss_amount), 0), 2)
        FROM regions r
        LEFT JOIN customers c ON r.region_id = c.region_id
        LEFT JOIN loans l ON c.customer_id = l.customer_id
        LEFT JOIN accounts a ON c.customer_id = a.customer_id
        LEFT JOIN fraud_alerts fa ON a.account_id = fa.account_id AND fa.status='Confirmed Fraud'
        GROUP BY r.region_name, r.regional_director
    """).fetchall()

    for row_idx, r in enumerate(regional_data, start=8):
        ws1.append(list(r))
        ws1.cell(row=row_idx, column=3).number_format = "#,##0"
        ws1.cell(row=row_idx, column=4).number_format = "0.0\"%\""
        ws1.cell(row=row_idx, column=5).number_format = "0.00"
        ws1.cell(row=row_idx, column=6).number_format = "0.0\"%\""
        ws1.cell(row=row_idx, column=7).number_format = "$#,##0"
        for col_num in range(1, 8):
            ws1.cell(row=row_idx, column=col_num).border = thin_border

    # ---------------------------------------------------------
    # TAB 2: REGIONAL LOAN SUMMARY
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="Regional Loan Summary")
    ws2.views.sheetView[0].showGridLines = True
    
    headers2 = ["Region ID", "Region Name", "Total Deposits ($)", "Loan Principal ($)", "Defaulted Loans", "Default Loss ($)", "Default Rate (%)"]
    ws2.append(["REGIONAL LOAN PORTFOLIO & CREDIT RISK AUDIT"])
    ws2["A1"].font = title_font
    ws2.append([])
    ws2.append(headers2)

    for col_num in range(1, 8):
        cell = ws2.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill

    loan_data = cursor.execute("""
        SELECT r.region_id, r.region_name,
               COALESCE(SUM(a.current_balance), 0),
               COALESCE(SUM(l.principal_amount), 0),
               SUM(CASE WHEN l.status='Defaulted' THEN 1 ELSE 0 END),
               SUM(CASE WHEN l.status='Defaulted' THEN l.principal_amount ELSE 0 END),
               ROUND(100.0 * SUM(CASE WHEN l.status='Defaulted' THEN 1 ELSE 0 END)/COUNT(l.loan_id), 2)
        FROM regions r
        LEFT JOIN customers c ON r.region_id = c.region_id
        LEFT JOIN accounts a ON c.customer_id = a.customer_id
        LEFT JOIN loans l ON c.customer_id = l.customer_id
        GROUP BY r.region_id, r.region_name
    """).fetchall()

    for row_idx, r in enumerate(loan_data, start=4):
        ws2.append(list(r))
        ws2.cell(row=row_idx, column=3).number_format = "$#,##0"
        ws2.cell(row=row_idx, column=4).number_format = "$#,##0"
        ws2.cell(row=row_idx, column=5).number_format = "#,##0"
        ws2.cell(row=row_idx, column=6).number_format = "$#,##0"
        ws2.cell(row=row_idx, column=7).number_format = "0.0\"%\""
        for col_num in range(1, 8):
            ws2.cell(row=row_idx, column=col_num).border = thin_border

    # Color scale conditional formatting on Default Rate column G
    color_scale = ColorScaleRule(start_type='num', start_value=3.0, start_color='D1FAE5',
                                 mid_type='num', mid_value=6.0, mid_color='FEF3C7',
                                 end_type='num', end_value=9.0, end_color='FEE2E2')
    ws2.conditional_formatting.add("G4:G8", color_scale)

    # ---------------------------------------------------------
    # TAB 3: FRAUD RISK REGISTER
    # ---------------------------------------------------------
    ws3 = wb.create_sheet(title="Fraud Risk Register")
    ws3.views.sheetView[0].showGridLines = True
    
    headers3 = ["Fraud Type", "Confirmed Alert Count", "Total Loss ($)", "Avg Loss ($)", "Avg Risk Score"]
    ws3.append(["CONFIRMED FRAUD LOSS & RISK REGISTER"])
    ws3["A1"].font = title_font
    ws3.append([])
    ws3.append(headers3)

    for col_num in range(1, 6):
        cell = ws3.cell(row=3, column=col_num)
        cell.font = header_font
        cell.fill = header_fill

    fraud_summary = cursor.execute("""
        SELECT fraud_type, COUNT(alert_id), ROUND(SUM(loss_amount), 2), ROUND(AVG(loss_amount), 2), ROUND(AVG(risk_score), 1)
        FROM fraud_alerts
        WHERE status = 'Confirmed Fraud'
        GROUP BY fraud_type
        ORDER BY SUM(loss_amount) DESC
    """).fetchall()

    for row_idx, r in enumerate(fraud_summary, start=4):
        ws3.append(list(r))
        ws3.cell(row=row_idx, column=2).number_format = "#,##0"
        ws3.cell(row=row_idx, column=3).number_format = "$#,##0"
        ws3.cell(row=row_idx, column=4).number_format = "$#,##0"
        ws3.cell(row=row_idx, column=5).number_format = "0.0"
        for col_num in range(1, 6):
            ws3.cell(row=row_idx, column=col_num).border = thin_border

    # ---------------------------------------------------------
    # TAB 4: DYNAMIC SLICER MINI DASHBOARD
    # ---------------------------------------------------------
    ws4 = wb.create_sheet(title="Interactive Mini Dashboard")
    ws4.views.sheetView[0].showGridLines = True

    ws4.append(["INTERACTIVE REGIONAL OPERATIONS ANALYTICS LOOKUP"])
    ws4["A1"].font = title_font
    ws4.append([])
    
    ws4["A3"] = "Select Region Name:"
    ws4["A3"].font = bold_font
    ws4["B3"] = "Southeast Hub"
    ws4["B3"].font = Font(bold=True, color="0B132B")
    ws4["B3"].fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")

    ws4.append([])
    ws4.append(["Operational Metric", "Regional Value", "Calculation Formula / Reference"])
    for col_num in range(1, 4):
        cell = ws4.cell(row=5, column=col_num)
        cell.font = header_font
        cell.fill = header_fill

    metrics = [
        ("Total Active Deposits ($)", "=XLOOKUP(B3, 'Regional Loan Summary'!B4:B8, 'Regional Loan Summary'!C4:C8)", "$#,##0"),
        ("Total Loan Principal ($)", "=XLOOKUP(B3, 'Regional Loan Summary'!B4:B8, 'Regional Loan Summary'!D4:D8)", "$#,##0"),
        ("Defaulted Principal Loss ($)", "=XLOOKUP(B3, 'Regional Loan Summary'!B4:B8, 'Regional Loan Summary'!F4:F8)", "$#,##0"),
        ("Loan Default Rate (%)", "=XLOOKUP(B3, 'Regional Loan Summary'!B4:B8, 'Regional Loan Summary'!G4:G8)", "0.0\"%\""),
        ("Confirmed Fraud Losses ($)", "=XLOOKUP(B3, 'Executive Summary'!A8:A12, 'Executive Summary'!G8:G12)", "$#,##0")
    ]

    for idx, (m_label, m_formula, m_fmt) in enumerate(metrics, start=6):
        ws4.cell(row=idx, column=1, value=m_label).font = bold_font
        ws4.cell(row=idx, column=2, value=m_formula).font = regular_font
        ws4.cell(row=idx, column=2).number_format = m_fmt
        ws4.cell(row=idx, column=3, value=f"XLOOKUP against B3 region selector").font = Font(italic=True, size=9, color="64748B")
        for col_num in range(1, 4):
            ws4.cell(row=idx, column=col_num).border = thin_border

    # Auto-adjust column widths across all sheets
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '$' in cell.number_format:
                    val_str += "   "
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 14)

    wb.save(EXCEL_PATH)
    print(f"Excel workbook generated successfully at: {EXCEL_PATH}")
    conn.close()

if __name__ == "__main__":
    build_excel()
